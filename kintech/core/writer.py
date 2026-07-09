from __future__ import annotations
import csv
import logging
from pathlib import Path
from typing import Optional
from .transform import OutputTable
log = logging.getLogger(__name__)
_SUPPORTED_EXTENSIONS = {'.csv', '.txt', '.xlsx'}

def write_output(table: OutputTable, out_path: Path, sheet_name: str='Data') -> None:
    ext = out_path.suffix.lower()
    if ext not in _SUPPORTED_EXTENSIONS:
        raise ValueError(f'Unsupported output extension {ext!r}. Supported: {sorted(_SUPPORTED_EXTENSIONS)}')
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if ext == '.csv':
        _write_delimited(table, out_path, delimiter=',')
    elif ext == '.txt':
        _write_delimited(table, out_path, delimiter='\t')
    elif ext == '.xlsx':
        _write_xlsx(table, out_path, sheet_name=sheet_name)

def _write_delimited(table: OutputTable, out_path: Path, delimiter: str) -> None:
    log.debug('Writing %s rows x %s cols to %s (delimiter=%r)', len(table.rows), len(table.columns), out_path, delimiter)
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerow(table.columns)
        for row in table.rows:
            writer.writerow([_format_value(v, col, table.decimals) for v, col in zip(row, table.columns)])

def _format_value(value, col_name: str, decimals: dict) -> str:
    if value is None:
        return ''
    if isinstance(value, float):
        d = decimals.get(col_name)
        if d is not None:
            return f'{value:.{d}f}'
        return _format_sigfigs(value, 5)
    return str(value)

def _format_sigfigs(x: float, sig: int) -> str:
    if x == 0:
        return '0.' + '0' * (sig - 1) if sig > 1 else '0'
    from math import floor, log10
    exponent = int(floor(log10(abs(x))))
    decimals = max(sig - exponent - 1, 0)
    return f'{x:.{decimals}f}'

def _write_xlsx(table: OutputTable, out_path: Path, sheet_name: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Writing .xlsx output requires the 'openpyxl' package. Install it with: pip install openpyxl --break-system-packages") from exc
    log.debug('Writing %s rows x %s cols to %s (xlsx, sheet=%r)', len(table.rows), len(table.columns), out_path, sheet_name)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(table.columns)
    for row in table.rows:
        formatted = [round(v, table.decimals[col]) if isinstance(v, float) and col in table.decimals else v for v, col in zip(row, table.columns)]
        ws.append(formatted)
    for idx, col_name in enumerate(table.columns, start=1):
        d = table.decimals.get(col_name)
        if d is not None:
            fmt = '0.' + '0' * d if d > 0 else '0'
            for row_idx in range(2, len(table.rows) + 2):
                ws.cell(row=row_idx, column=idx).number_format = fmt
    ws.freeze_panes = 'A2'
    for idx, col_name in enumerate(table.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(r[idx - 1])) for r in table.rows[:500] if r[idx - 1] is not None] or [10])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 28)
    wb.save(out_path)